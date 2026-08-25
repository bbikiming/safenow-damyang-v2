#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
분류기준 v3.3 생성기 — 담양군_2025_문서분류_대시보드.xlsx 「기준표」 → js/doc-taxonomy-data.js

  무엇이 달라지나 (v3 CSV 177단계 → v3.3 엑셀 213단계)
  ────────────────────────────────────────────────────
  · 단계 **삭제 0건 · 추가 36건**. 현행 177 코드가 전부 그대로 있어 기존 문서 매핑
    (stageIds 는 코드다)이 무수정으로 살아난다.
  · 「마련」과 「점검」이 **단계로 쪼개졌다** — 종전 «매뉴얼 마련 및 점검»(반기 1회)이
    «대응 매뉴얼 마련»(EVENT)과 «조치 여부 반기 점검»(HALF, 신설)으로 갈렸다.
    2026-08-21 회의의 «제도는 마련했으나 실행 기록이 없음»이 이 분리로 표현된다.
    그래서 이름이 바뀐 13건은 전부 «및 점검»·«및 반기 평가» 부분이 떨어져 나간 것이다.
  · 그 결과 **법정주기가 14건 바뀐다** — 9건이 「반기/매월/연 1회」 → 「정기주기 없음」.
    마련은 한 번 하면 되는 일이므로 맞는 방향이고, 대신 점검이 새 단계로 독립했다.

  설계 원칙 — 「신규가 주는 것은 덮어쓰고, 신규에 없는 우리 판단은 이어받는다」
  ──────────────────────────────────────────────────────────────────
  엑셀 기준표 13열은 **발주처·컨설팅이 확정한 사실**이라 그대로 싣는다.
  반면 `수행경로`·`완료판정`·`분류확신도` 같은 열은 **우리가 만든 판단**이고 엑셀에
  없다. 지우면 화면의 «어디서 수행하나»가 통째로 사라지므로, 현행 CSV 에서 **코드로**
  이어받는다(이름은 13건이 바뀌었으니 이름으로 이으면 끊긴다).
  36개 신규 단계는 그 값이 비어 있다 — 지어내지 않는다.

  새로 들어오는 축 5개 — 지금까지 «자료 미취합»으로 비워 두던 자리들이다
  ──────────────────────────────────────────────────────────────
  · 이행단위   → levelSrc  L1 군 103 · L2 부서 68 · L3 시설 29 · L3 공사·용역 13
                 (우리 `적용수준` 열이 177행 전부 빈값이라 문자열 파생 추정 중이었다)
  · 주기코드   → cycleCode 명시값. 문자열 파싱보다 **우선**한다.
                 ⚠ GRADE·TERM·MULTIYEAR·BIENNIAL 4종은 연 단위 고정 회차가 아니다.
  · 이행률 포함 → inRate    Y 72 · N 141. **이행률의 분모가 213 이 아니라 72 다.**
  · 이행의무 유형 → dutyKind 정기주기 70 · 상시·최초 52 · 조건부 88 · 산출불가 3
  · 조문상 의무성 → dutyBasis 의무 157 · 미검증 33 · 해당없음 9 · 재량 2 …

  검증 — 엑셀이 정답지를 갖고 있다
  ──────────────────────────────
  「이행률」 시트가 L1 97% · L2 63% · L3 시설 100% · L3 공사 0% · 합계 82%(59/72)를
  이미 계산해 두었다. 우리가 같은 값을 재현하면 그릇이 맞는 것이다. 이 스크립트는
  그 재현을 검사하고 어긋나면 파일을 쓰지 않는다.

  실행: python3 tools/build-taxonomy-v33.py --xlsx <경로> [--carry data/…v3.csv]
"""

import argparse
import csv
import io
import json
import os
import re
import sys
from collections import Counter, OrderedDict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'js', 'doc-taxonomy-data.js')

# 「이행단위」 원문 → (계층, 하위 구분)
UNIT = {
    'L1 군(관리주체)':        ('L1', ''),
    'L2 부서':               ('L2', ''),
    'L3 관리대상(시설)':      ('L3', 'facility'),
    'L3 관리대상(공사·용역)': ('L3', 'work'),
}

# 우리가 만든 판단 — 엑셀에 없어 현행 CSV 에서 코드로 이어받는다
CARRY = ['적용대상', '이행주체', '증빙문서 예시(2025년 실제 문서명)',
         '재난안전과 운영주기(참고)', '수행경로', '완료판정', '분류확신도', '확인필요사유']


def norm(v):
    return '' if v is None else str(v).strip()


def load_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['기준표']
    hdr, rows = None, []
    for r in ws.iter_rows(min_row=1, values_only=True):
        if hdr is None:
            if norm(r[0]) == '업무단계코드':
                hdr = [norm(x) for x in r]
            continue
        if norm(r[0]):
            rows.append(r)
    idx = {h: i for i, h in enumerate(hdr) if h}
    out = [{h: norm(r[idx[h]]) for h in idx} for r in rows]

    # ── 「EVENT생성방식」 시트 — 연계 개발자 회신이 반영된 **확정 설계** ──────────
    # EVENT 주기 140단계가 «언제 어떻게 생기는가»를 정한다. 우리가 종전에
    # «어느 결과를 이행으로 볼지 정해지지 않았습니다» 라고 뭉뚱그리던 자리에,
    # 이 시트는 **연계 시스템·인터페이스 번호·미확보 항목·조치**까지 갖고 있다.
    # 판정 D(보류) 19건은 전부 계약정보시스템(차세대 e호조) 제공 항목 미수령이다.
    gen = {}
    if 'EVENT생성방식' in wb.sheetnames:
        wsg = wb['EVENT생성방식']
        ghdr = None
        for r in wsg.iter_rows(min_row=1, values_only=True):
            if ghdr is None:
                if norm(r[0]) == '업무단계코드':
                    ghdr = [norm(x) for x in r]
                continue
            if not norm(r[0]):
                continue
            gi = {h: k for k, h in enumerate(ghdr) if h}
            gen[norm(r[0])] = {h: norm(r[gi[h]]) for h in gi}

    # 정답지 — 「이행률」 시트의 이행단위별 집계
    truth = {}
    ws2 = wb['이행률']
    seen_hdr = False
    for r in ws2.iter_rows(min_row=1, values_only=True):
        a = norm(r[0])
        if a == '이행단위':
            seen_hdr = True
            continue
        if seen_hdr and a in UNIT:
            truth[a] = {'total': int(r[1]), 'done': int(r[2]), 'miss': int(r[3])}
        elif seen_hdr and a == '합계':
            truth['합계'] = {'total': int(r[1]), 'done': int(r[2]), 'miss': int(r[3])}
            break

    # 문서목록 — 단계명별 건수(정답지 재현에 쓴다)
    ws3 = wb['문서목록']
    h3, docs = None, Counter()
    for r in ws3.iter_rows(min_row=4, values_only=True):
        if h3 is None:
            h3 = [norm(x) for x in r]
            continue
        if r[0] is None and r[1] is None:
            continue
        j = {h: k for k, h in enumerate(h3) if h}
        nm = norm(r[j['업무단계명 (수정 가능)']])
        if nm:
            docs[nm] += 1
    return out, truth, docs, gen


def load_carry(path):
    if not path or not os.path.exists(path):
        return {}
    rows = list(csv.DictReader(io.open(path, encoding='utf-8-sig')))
    return {norm(r.get('업무단계코드')): r for r in rows}


def item_of(code):
    """업무단계코드 → 이행항목코드. CIT-01-01 → CIT-01 (앞 두 마디)."""
    p = code.split('-')
    return '-'.join(p[:2]) if len(p) >= 2 else code


def parse_paths(s):
    """'PROGRAM:RSK_REGULAR' · 'ELECTRONIC_DOC' · 'ATTACHMENT' → paths[]"""
    s = norm(s)
    if not s:
        return []
    out = []
    for tok in [t.strip() for t in s.split('|') if t.strip()]:
        kind, _, code = tok.partition(':')
        out.append({'type': kind.strip().upper(), 'code': code.strip()})
    return out


def parse_done(s):
    s = norm(s)
    if not s:
        return {'kind': 'DOC_COUNT', 'key': ''}
    kind, _, key = s.partition(':')
    return {'kind': kind.strip().upper(), 'key': key.strip()}


def build(xrows, carry, gen):
    stages, items = [], OrderedDict()
    for r in xrows:
        code = r['업무단계코드']
        iid = item_of(code)
        old = carry.get(code, {})
        lvl, kind = UNIT.get(r.get('이행단위', ''), ('', ''))

        s = {
            'id': code,
            'itemId': iid,
            'name': r['하위 업무단계명'],
            'law': r['법령근거'],
            'legalCycle': r['법정주기'],
            'opCycle': norm(old.get('재난안전과 운영주기(참고)')),
            'timing': r['수행시점조건'],

            # ── 엑셀이 확정해 준 축 (v3.3 신규) ──────────────────────────
            'cycleCode': r.get('주기코드', ''),      # 문자열 파싱보다 우선한다
            'levelSrc': lvl,                          # 계층 «추정»이 «확정»이 된다
            'unitKind': kind,                         # L3 안의 시설 / 공사·용역
            'inRate': (r.get('이행률 포함', '') == 'Y'),   # 이행률 분모 여부
            'dutyKind': r.get('이행의무 유형', ''),
            'dutyBasis': r.get('조문상 의무성', ''),
            'deliveryKind': r.get('이행 방식 유형', ''),
            'eventGen': r.get('EVENT 생성방식', ''),

            # ── 우리 판단 — 코드로 이어받는다(신규 36건은 빈다) ──────────
            'paths': parse_paths(old.get('수행경로')),
            'doneRule': parse_done(old.get('완료판정')),
            'typeConf': (norm(old.get('분류확신도')) or 'UNKNOWN').upper(),
            'typeNote': norm(old.get('확인필요사유')),
            'target': norm(old.get('적용대상')),
            'actor': norm(old.get('이행주체')),
            'ex': norm(old.get('증빙문서 예시(2025년 실제 문서명)')),
        }
        s['taskType'] = s['paths'][0]['type'] if s['paths'] else 'UNKNOWN'

        # ── EVENT 생성방식 확정 (연계 개발자 회신) ──────────────────────────
        # 값이 없는 단계(EVENT 주기가 아닌 73건)는 필드 자체를 만들지 않는다.
        g = gen.get(code)
        if g:
            link = {
                'gen': g.get('최종 생성 방식', ''),
                'sys': g.get('연계 시스템', ''),
                'iface': g.get('연동 인터페이스', ''),
                'grade': g.get('연계 실현 판정', ''),      # A 확보 · B 일부 부족 · C · D 보류
                'missing': g.get('미확보 항목(개발자 회신)', ''),
                'why': g.get('판정 사유', ''),
                'action': g.get('조치', ''),
                'check': g.get('연동 확인 필요', ''),
            }
            if any(link.values()):
                s['link'] = {k: v for k, v in link.items() if v}
        stages.append(s)

        it = items.setdefault(iid, {'id': iid, 'name': r['법정 이행항목명'],
                                    'stageIds': [], 'lawBases': [], 'targets': [],
                                    'actors': [], 'hasCycle': False})
        it['stageIds'].append(code)
        for k, v in (('lawBases', s['law']), ('targets', s['target']), ('actors', s['actor'])):
            if v and v not in it[k]:
                it[k].append(v)
        if s['cycleCode'] != 'EVENT':
            it['hasCycle'] = True
    return stages, list(items.values())


def verify(stages, truth, docs):
    """엑셀 「이행률」 시트를 재현한다 — 못 하면 그릇이 틀린 것이다."""
    byname = {s['name']: s for s in stages}
    errs, got = [], {}
    for s in stages:
        if not s['inRate']:
            continue
        unit = next((k for k, v in UNIT.items() if v == (s['levelSrc'], s['unitKind'])), None)
        if not unit:
            errs.append('이행단위를 못 읽음: ' + s['id'])
            continue
        g = got.setdefault(unit, {'total': 0, 'done': 0, 'miss': 0})
        g['total'] += 1
        if docs.get(s['name'], 0) > 0:
            g['done'] += 1
        else:
            g['miss'] += 1
    tot = {'total': 0, 'done': 0, 'miss': 0}
    for v in got.values():
        for k in tot:
            tot[k] += v[k]
    got['합계'] = tot

    print('\n=== 이행률 재현 (엑셀 「이행률」 시트가 정답지) ===')
    for k in list(UNIT) + ['합계']:
        w, g = truth.get(k), got.get(k)
        if not w:
            continue
        ok = g and all(g[x] == w[x] for x in ('total', 'done', 'miss'))
        print('  %-22s 정답 %3d/%3d/%3d · 재현 %s  %s' % (
            k, w['total'], w['done'], w['miss'],
            ('%3d/%3d/%3d' % (g['total'], g['done'], g['miss'])) if g else '(없음)',
            '✓' if ok else '✗'))
        if not ok:
            errs.append('%s 불일치' % k)
    return errs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--carry', default=os.path.join(ROOT, 'data', '담양군_안전보건_문서분류기준_v3.csv'))
    ap.add_argument('--force', action='store_true', help='검증 실패해도 쓴다(권장하지 않음)')
    a = ap.parse_args()

    xrows, truth, docs, gen = load_xlsx(a.xlsx)
    carry = load_carry(a.carry)
    stages, items = build(xrows, carry, gen)

    print('이행항목 %d · 업무단계 %d' % (len(items), len(stages)))
    print('  이행단위 :', dict(Counter(s['levelSrc'] + (('/' + s['unitKind']) if s['unitKind'] else '') for s in stages)))
    print('  주기코드 :', dict(Counter(s['cycleCode'] for s in stages).most_common()))
    print('  이행률 포함 Y: %d / %d' % (sum(1 for s in stages if s['inRate']), len(stages)))
    lk = [s for s in stages if s.get('link')]
    print('  EVENT 생성방식: %d단계' % len(lk),
          dict(Counter(s['link'].get('gen', '') for s in lk).most_common()))
    print('  연계 실현 판정 :', dict(Counter(s['link'].get('grade', '—') for s in lk).most_common()))
    print('  연동 인터페이스:', dict(Counter(s['link'].get('iface', '(없음)') for s in lk).most_common()))
    print('  우리 판단 이어받음: 수행경로 %d · 완료판정 %d · 적용대상 %d'
          % (sum(1 for s in stages if s['paths']),
             sum(1 for s in stages if s['doneRule']['kind'] == 'PROBE'),
             sum(1 for s in stages if s['target'])))

    errs = verify(stages, truth, docs)
    if errs and not a.force:
        print('\n검증 실패 — 파일을 쓰지 않습니다:')
        for e in sorted(set(errs))[:10]:
            print('  ·', e)
        sys.exit(1)

    meta = {'source': os.path.basename(a.xlsx), 'sheet': '기준표', 'version': 'v3.3',
            'items': len(items), 'stages': len(stages),
            'carriedFrom': os.path.basename(a.carry) if carry else '',
            'rateDenominator': sum(1 for s in stages if s['inRate'])}
    hdr = '''/* =========================================================================
 * 업무문서 분류 v3.3 — 법정 이행항목 %d · 하위 업무단계 %d (DYDOCT)
 *   ※ 생성물 — 손으로 고치지 말 것. 원본 엑셀을 고치고 재생성한다.
 *      생성기: tools/build-taxonomy-v33.py
 *      원본:   %s 「기준표」
 *
 *   [v3 → v3.3] 단계 삭제 0 · 추가 36. 「마련」과 「점검」이 단계로 쪼개졌다 —
 *   «매뉴얼 마련 및 점검»(반기)이 «마련»(EVENT)과 «조치 여부 반기 점검»(신설)으로
 *   갈렸다. 2026-08-21 회의의 «제도는 마련했으나 실행 기록이 없음»이 이 분리다.
 *
 *   [새 축] cycleCode(주기 명시값·파싱보다 우선) · levelSrc/unitKind(계층 확정값) ·
 *   inRate(이행률 분모 여부 — 분모는 %d 이지 %d 가 아니다) · dutyKind · dutyBasis ·
 *   deliveryKind · eventGen.
 *
 *   [이어받은 값] paths·doneRule·typeConf·typeNote·target·actor·ex 는 엑셀에 없는
 *   **우리 판단**이라 v3 CSV 에서 코드로 이어받았다. 신규 36단계는 비어 있다.
 * ========================================================================= */
''' % (len(items), len(stages), os.path.basename(a.xlsx), meta['rateDenominator'], len(stages))

    body = 'window.DYDOCT = {\n  META: ' + json.dumps(meta, ensure_ascii=False, indent=2) + ',\n'
    body += '  ITEMS: [\n' + ',\n'.join('    ' + json.dumps(x, ensure_ascii=False) for x in items) + '\n  ],\n'
    body += '  STAGES: [\n' + ',\n'.join('    ' + json.dumps(x, ensure_ascii=False) for x in stages) + '\n  ]\n};\n'
    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(hdr + body)
    print('\n생성: %s (%.1f KB)' % (os.path.relpath(OUT, ROOT), os.path.getsize(OUT) / 1024))


if __name__ == '__main__':
    main()
